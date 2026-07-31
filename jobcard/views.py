import profile

import jobcard

from .decorators import role_required
import requests
from io import BytesIO

from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.contrib import messages
from .forms import TempSubmissionForm, JobCardForm, JobCardPrepopulateForm
from .models import TempSubmission, ShiftSubmission, JobCard, LINE_CHOICES, ActiveShift
from datetime import timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from django.utils.timezone import now
from .models import ActiveShift

# Active shift / current date
def ensure_active_shift_is_current(active):
    current_time = timezone.localtime()

    # ✅ Get correct production date based on shift
    correct_date = get_production_date(active.shift, current_time)

    # Only update if needed
    if active.date != correct_date:
        active.date = correct_date
        active.last_reset = timezone.now()
        active.event_type = "shift_start"
        active.line = None
        active.save()

    return active


# -----------------------------
# Helper function (kept for fallback safety only)
# -----------------------------
def get_production_date(shift: str, current_time=None):
    now = current_time or timezone.localtime()
    today = now.date()
    if shift.lower() == "night":
        cutoff = time(5, 30)
        if now.time() < cutoff:
            return today - timedelta(days=1)
    return today

# -----------------------------
# EXCEL EXPORT (OpenPyXL)
# -----------------------------
def export_jobcards_csv(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    line = request.GET.get("line")
    shift = request.GET.get("shift")

    jobcards = JobCard.objects.all()
    if start_date and end_date:
        jobcards = jobcards.filter(
            date__range=[start_date, end_date]
        )


    if line:
        jobcards = jobcards.filter(line=line)

    if shift:
        jobcards = jobcards.filter(shift=shift)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "JobCards"

    # Cache downloaded signatures
    signature_cache = {}

    header = [
        'Date','Line','Shift','WO Number','Product Code','Product Name','Target Quantity',
        'Hour1','Hour2','Hour3','Hour4','Hour5','Hour6','Hour7','Hour8','Hour9','Hour10','Hour11','Hour12',
        'Total Output',
        'Jar Damage','Cap Damage','Front Label Damage','Back Label Damage','Carton Damage',
        'Sleeve Damage','Sticker Damage','Tube Damage','Packets Damage','Roll On Ball Damage','Jar Pump Damage',
        'Total Damage',
        'Jar Reject','Cap Reject','Front Label Reject','Back Label Reject','Carton Reject',
        'Sleeve Reject','Sticker Reject','Tube Reject','Packets Reject','Roll On Ball Reject','Jar Pump Reject',
        'Total Reject',
        'Operators','Supervisors',
        'Supervisor Signature',
        'Line Captain Signature'
    ]

    sheet.append(header)

    # Header styling
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 2

    for jc in jobcards:

        row = [
            jc.date,
            jc.line,
            jc.shift,
            jc.wo_number,
            jc.product_code,
            jc.product_name,
            jc.target_quantity,

            jc.hour1,
            jc.hour2,
            jc.hour3,
            jc.hour4,
            jc.hour5,
            jc.hour6,
            jc.hour7,
            jc.hour8,
            jc.hour9,
            jc.hour10,
            jc.hour11,
            jc.hour12,

            jc.total_output(),

            jc.jar_damage,
            jc.cap_damage,
            jc.front_label_damage,
            jc.back_label_damage,
            jc.carton_damage,
            jc.sleeve_damage,
            jc.sticker_damage,
            jc.tube_damage,
            jc.packets_damage,
            jc.roll_on_ball_damage,
            jc.jar_pump_damage,

            jc.total_damage(),

            jc.jar_reject,
            jc.cap_reject,
            jc.front_label_reject,
            jc.back_label_reject,
            jc.carton_reject,
            jc.sleeve_reject,
            jc.sticker_reject,
            jc.tube_reject,
            jc.packets_reject,
            jc.roll_on_ball_reject,
            jc.jar_pump_reject,

            jc.total_reject(),

            jc.operator_names,
            jc.supervisor_names,

            "",   # Supervisor Signature placeholder
            ""    # Line Captain Signature placeholder
        ]

        sheet.append(row)

        # -------------------------
        # Supervisor Signature
        # -------------------------
        if jc.supervisor_signature:
            try:
                url = jc.supervisor_signature.url

                if url not in signature_cache:

                    response = requests.get(url, timeout=2)

                    if response.status_code == 200:
                        signature_cache[url] = response.content
                    else:
                        signature_cache[url] = None

                if signature_cache[url]:

                    img = Image(BytesIO(signature_cache[url]))
                    img.width = 90
                    img.height = 45

                    sheet.add_image(img, f"AW{current_row}")
                    sheet.row_dimensions[current_row].height = 40

                else:
                    sheet[f"AW{current_row}"] = "Signature Missing"

            except Exception:
                sheet[f"AW{current_row}"] = "Signature Missing"
        # -------------------------
        # Line Captain Signature
        # -------------------------
        if jc.line_captain_signature:
            try:
                url = jc.line_captain_signature.url

                if url not in signature_cache:

                    response = requests.get(url, timeout=2)

                    if response.status_code == 200:
                        signature_cache[url] = response.content
                    else:
                        signature_cache[url] = None

                if signature_cache[url]:

                    img = Image(BytesIO(signature_cache[url]))
                    img.width = 90
                    img.height = 45

                    sheet.add_image(img, f"AX{current_row}")
                    sheet.row_dimensions[current_row].height = 40

                else:
                    sheet[f"AX{current_row}"] = "Signature Missing"

            except Exception:
                sheet[f"AX{current_row}"] = "Signature Missing"

        current_row += 1

    # Auto-size columns
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 3, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        f'attachment; filename="jobcards_{start_date}_to_{end_date}.xlsx"'
    )

    workbook.save(response)

    return response

# -----------------------------
# TEMP SUBMISSION (FIXED SHIFT SOURCE ONLY)
# -----------------------------
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction, IntegrityError

@csrf_exempt
def temp_submission(request):
    active = ActiveShift.objects.first()

    if not active:
        active = ActiveShift.objects.create(
            shift="Day",
            date=timezone.localdate()
        )

    active = ensure_active_shift_is_current(active)

    shift = active.shift
    target_date = active.date

    selected_line = request.GET.get("line")
    lines = [l[0] for l in LINE_CHOICES]
    forms_data = []

    # ✅ SAFE OPERATOR FIX (NO CRASH)
    operator = request.user if request.user.is_authenticated else None

    # ---------------- AJAX SAVE ----------------
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        line = request.POST.get("line")

        try:
            # ✅ RETRY SAFE BLOCK
            for _ in range(3):
                try:
                    with transaction.atomic():
                        obj, _ = TempSubmission.objects.select_for_update().get_or_create(
                            operator=operator,
                            date=target_date,
                            shift=shift,
                            line=line
                        )
                    break
                except IntegrityError:
                    continue
            else:
                return JsonResponse({"error": "Database conflict, retry"}, status=500)

            updated_fields = []

            for i in range(1, 13):
                field = f"hour{i}"
                new_val = request.POST.get(field)
                old_val = getattr(obj, field)

                if new_val in [None, ""]:
                    continue

                try:
                    new_val = float(new_val)
                except:
                    continue

                # ✅ LOCK FIELD (unchanged logic)
                if old_val not in [None, 0, 0.0]:
                    return JsonResponse(
                        {"error": f"{field.upper()} already submitted and locked."},
                        status=403
                    )

                if new_val == 0:
                    continue

                setattr(obj, field, new_val)
                updated_fields.append(i)

            obj.save()

            return JsonResponse({"success": True, "updated": updated_fields})

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)

    # ---------------- FORM LOAD ----------------
    for line in lines:
        if selected_line and line != selected_line:
            continue

        try:
            # ✅ RETRY SAFE BLOCK
            for _ in range(3):
                try:
                    with transaction.atomic():
                        obj, created = TempSubmission.objects.get_or_create(
                            operator=operator,
                            date=target_date,
                            shift=shift,
                            line=line
                        )
                    break
                except IntegrityError:
                    continue
            else:
                obj = TempSubmission.objects.filter(
                    operator=operator,
                    date=target_date,
                    shift=shift,
                    line=line
                ).first()
                created = False

            if created:
                for i in range(1, 13):
                    setattr(obj, f"hour{i}", 0)
                obj.save()

        except Exception:
            obj = TempSubmission.objects.filter(
                operator=operator,
                date=target_date,
                shift=shift,
                line=line
            ).first()

        form = TempSubmissionForm(instance=obj)
        forms_data.append((line, form, obj))

    return render(request, "temp_submission_form.html", {
        "forms_data": forms_data,
        "shift": shift,
        "selected_line": selected_line
    })
# -----------------------------
# SUPERVISOR DASHBOARD (FIXED - STRICT ACTIVE SHIFT ONLY)
# -----------------------------
@role_required("supervisor")
def supervisor_dashboard(request):
    active = ActiveShift.objects.first()

    if not active:
        active = ActiveShift.objects.create(
            shift="Day",
            date=timezone.localdate()
        )

    active = ensure_active_shift_is_current(active)

    # ✅ FORCE SYSTEM SINGLE SOURCE OF TRUTH
    shift = active.shift
    target_date = active.date

    submissions = TempSubmission.objects.filter(
        date=target_date,
        shift=shift
    ).order_by("line", "operator")

    lines = [l[0] for l in LINE_CHOICES]
    global_locked_hours = []

    for h in range(1, 13):
        filled_lines = submissions.exclude(**{f"hour{h}__isnull": True}) \
                                  .exclude(**{f"hour{h}": 0}) \
                                  .values("line").distinct().count()
        if filled_lines >= len(lines):
            global_locked_hours.append(h)

    dashboard_data = {}
    for sub in submissions:
        key = f"{sub.line}_{sub.shift}"
        if key not in dashboard_data:
            dashboard_data[key] = {
                "submissions": [],
                "hour_totals": [0]*12,
                "total": 0
            }

        dashboard_data[key]["submissions"].append(sub)

        hours = [getattr(sub, f"hour{i}") or 0 for i in range(1, 13)]
        for i in range(12):
            dashboard_data[key]["hour_totals"][i] += hours[i]

        dashboard_data[key]["total"] += sum(hours)

    # AJAX response
    if request.GET.get("ajax") == "1":
        clean_data = {}
        for key, data in dashboard_data.items():
            clean_data[key] = {
                "hour_totals": data["hour_totals"],
                "total": data["total"]
            }

        return JsonResponse({
            "dashboard_data": clean_data,
            "global_locked_hours": global_locked_hours
        }, safe=True)

    return render(request, "supervisor_dashboard.html", {
        "dashboard_data": dashboard_data,
        "today": target_date,
        "hour_range": range(1, 13),
        "shift": shift
    })

# -----------------------------
# RESET SHIFT (unchanged)
# -----------------------------
@role_required("supervisor")
def reset_shift(request):
    if request.method == "POST":
        shift = request.POST.get("shift")
        line = request.POST.get("line")

        active = ActiveShift.objects.first()

        if not active:
            active = ActiveShift.objects.create(
                shift=shift,
                date=timezone.localdate()
            )

        temp_query = TempSubmission.objects.filter(
            shift=shift,
            date=active.date
        )

        if line:
            temp_query = temp_query.filter(line=line)
            temp_query.delete()

            messages.success(
                request,
                f"✅ {shift} shift for line {line} has been reset successfully."
            )

            # ✅ LINE RESET EVENT
            active.event_type = "line_reset"
            active.line = line

        else:
            temp_query.delete()

            messages.success(
                request,
                f"✅ All lines for {shift} shift have been reset successfully."
            )

            # ✅ SHIFT RESET EVENT
            active.event_type = "shift_start"
            active.line = None

        # ✅ COMMON UPDATE (CRITICAL FIX)
        active.shift = shift
        active.last_reset = timezone.now()

        # 🔥 IMPORTANT: ensure event_type is never null
        if not active.event_type:
            active.event_type = "shift_start" if not line else "line_reset"

        active.save()

    return redirect("jobcard:supervisor_dashboard")

# -----------------------------
# FINALIZE SHIFT (unchanged)
# -----------------------------
def finalize_shift(request, line, shift):
    active = ActiveShift.objects.first()
    target_date = active.date if active else timezone.localdate()
    submissions = TempSubmission.objects.filter(
        date=target_date,
        line=line,
        shift=shift
    )
    aggregated_data = [{
        "operator": s.operator.username if s.operator else "Anonymous",
        "hours": [getattr(s, f"hour{i}") or 0 for i in range(1,13)],
        "total": s.total_output()
    } for s in submissions]
    shift_submission, created = ShiftSubmission.objects.get_or_create(
        date=target_date,
        line=line,
        shift=shift,
        defaults={"aggregated_data": aggregated_data}
    )
    if not created:
        shift_submission.aggregated_data = aggregated_data
        shift_submission.save()
    return redirect("jobcard:supervisor_dashboard")

# -----------------------------
# JOBCARD OPERATOR ENTRY (STRICT SHIFT MATCH)
# -----------------------------
@role_required("operator")
def jobcard_operator_entry(request):
    active = ActiveShift.objects.first()

    if not active:
        active = ActiveShift.objects.create(
            shift="Day",
            date=timezone.localdate()
        )

    active = ensure_active_shift_is_current(active)

    shift = active.shift
    jobcard_date = active.date
    line = request.POST.get("line") or request.GET.get("line")

    if not line:
        messages.warning(request, "Please select a Line first.")
        form = JobCardForm()
        return render(request, "jobcard_form.html", {
            "form": form,
            "shift": shift,
            "line": line
        })

    if request.method == "POST":
        wo_number = request.POST.get("wo_number")

        existing = JobCard.objects.filter(
            date=jobcard_date,
            line=line,
            shift=shift,
            wo_number=wo_number
        ).first()

        if existing:
            form = JobCardForm(request.POST, instance=existing)
        else:
            form = JobCardForm(request.POST)
            form.instance.date = jobcard_date
            form.instance.line = line
            form.instance.shift = shift

        if form.is_valid():
            jobcard = form.save(commit=False)
            profile = request.user.userprofile
            jobcard.line_captain_signature = profile.signature
            jobcard.date = jobcard_date
            jobcard.line = line
            jobcard.shift = shift
            jobcard.is_submitted = True
            jobcard.save()

            if not existing:
                temp_data = TempSubmission.objects.filter(
                    date=jobcard_date,
                    line=line,
                    shift=shift
                ).first()
                if temp_data:
                    for i in range(1, 13):
                        setattr(jobcard, f"hour{i}", getattr(temp_data, f"hour{i}", 0))
                    jobcard.save()

            messages.success(request, "✅ JobCard submitted successfully!")
            return redirect("jobcard:jobcard_success")
        else:
            messages.error(request, f"Errors: {form.errors}")

    else:
        form = JobCardForm()

    return render(request, "jobcard_form.html", {
        "form": form,
        "shift": shift,
        "line": line
    })

# -----------------------------
# JOBCARD SUCCESS (unchanged)
# -----------------------------
def jobcard_success(request):
    return render(request, "success.html")

# -----------------------------
# JOBCARD PREPOPULATE (unchanged logic)
# -----------------------------
@role_required("supervisor")
def jobcard_prepopulate(request):
    active = ActiveShift.objects.first()

    if not active:
        active = ActiveShift.objects.create(
            shift="Day",
            date=timezone.localdate()
        )

    active = ensure_active_shift_is_current(active)

    shift = active.shift
    jobcard_date = active.date

    if request.method == "POST":
        form = JobCardPrepopulateForm(request.POST)
        if form.is_valid():
            line = form.cleaned_data['line']
            profile = request.user.userprofile
            wo_number = form.cleaned_data.get('wo_number')

            jobcard, created = JobCard.objects.get_or_create(
                date=jobcard_date,
                line=line,
                shift=shift,
                wo_number=wo_number,
                defaults={
    **form.cleaned_data,
    "supervisor_signature": profile.signature
}
            )

            if not created:
                for field, value in form.cleaned_data.items():
                    if field not in ['line', 'wo_number']:
                        setattr(jobcard, field, value)

                # Copy supervisor's saved signature
                jobcard.supervisor_signature = profile.signature

                jobcard.save()
                messages.success(request, f"JobCard for {line} WO {wo_number} ({shift}) updated.")
            else:
                messages.success(request, f"JobCard for {line} WO {wo_number} ({shift}) created.")

            return redirect('jobcard:jobcard_prepopulate')
    else:
        form = JobCardPrepopulateForm()

    return render(request, "jobcard_prepopulate.html", {"form": form})

# -----------------------------
# GET JOBCARD AJAX (STRICT SHIFT MATCH)
# -----------------------------
def get_jobcard(request):
    line = request.GET.get("line")
    active = ActiveShift.objects.first()

    if not active:
        return JsonResponse({"error": "No active shift set. Please wait for supervisor to start a shift."})

    shift = active.shift
    target_date = active.date

    job = JobCard.objects.filter(
        line=line,
        shift=shift,
        date=target_date
    ).order_by('-id').first()

    if not job:
        return JsonResponse({"error": "No JobCard found for this line & shift."})

    temp = TempSubmission.objects.filter(
        date=target_date,
        line=line,
        shift=shift
    ).first()

    hours = []
    for i in range(1, 13):
        if temp and getattr(temp, f"hour{i}", None) is not None:
            hours.append(getattr(temp, f"hour{i}"))
        else:
            hours.append(getattr(job, f"hour{i}", 0))

    return JsonResponse({
        "wo_number": job.wo_number,
        "product_code": job.product_code,
        "product_name": job.product_name,
        "target_quantity": job.target_quantity,
        "operator_names": job.operator_names,
        "supervisor_names": job.supervisor_names,
        "supervisor_signature":
    job.supervisor_signature.url
    if job.supervisor_signature
    else "",
        "hours": hours,
        "submitted": bool(job.is_submitted)
    })

# -----------------------------
# CSRF FAILURE (unchanged)
# -----------------------------

def custom_400(request, exception):
    return render(request, "errors/400.html", status=400)


def custom_403(request, exception):
    return render(request, "errors/403.html", status=403)


def custom_404(request, exception):
    return render(request, "errors/404.html", status=404)


def custom_500(request):
    return render(request, "errors/500.html", status=500)

# One more view to provide active shift info for frontend
def get_active_shift(request):
    active = ActiveShift.objects.first()

    if not active:
        return JsonResponse({"error": "No active shift"})

    return JsonResponse({
        "shift": active.shift,
        "date": str(active.date),
        "last_reset": str(active.last_reset) if active.last_reset else None,
        "line": getattr(active, "line", None),
        "event_type": getattr(active, "event_type", None),
    })


# ACTIVE SHIFT SETTER (for supervisor to change shift)
def set_active_shift(request):
    if request.method == "POST":
        shift = request.POST.get("shift")

        active = ActiveShift.objects.first()
        if not active:
            active = ActiveShift.objects.create(
                shift=shift,
                date=timezone.localdate()
            )
        else:
            active.shift = shift
            active.save()

    return redirect("jobcard:supervisor_dashboard")


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password

from .models import UserProfile
from .forms import UserCreateForm


# =====================================
# ROLE DECORATOR (AUTHORIZATION LAYER)
# =====================================
from django.views.decorators.cache import never_cache

def role_required(role):
    def decorator(view_func):

        @never_cache
        @login_required(login_url="/jobcard/login/")
        def wrapper(request, *args, **kwargs):

            profile = UserProfile.objects.filter(user=request.user).first()

            if not profile:
                return redirect("/jobcard/login/")

            if profile.role.strip().lower() != role.strip().lower():
                return redirect("/jobcard/login/")

            return view_func(request, *args, **kwargs)

        return wrapper

    return decorator


# =====================================
# DASHBOARD (ROLE VIEW SWITCHER)
# =====================================
@login_required(login_url="/jobcard/login/")
def dashboard_home(request):

    profile = UserProfile.objects.filter(user=request.user).first()

    if not profile:
        return render(request, "auth_required.html", status=403)

    return render(request, "dashboard.html", {
        "role": profile.role,
        "is_developer": profile.role == "developer",
        "is_supervisor": profile.role == "supervisor",
        "is_operator": profile.role == "operator",
    })


# =====================================
# ROLE REDIRECT (AFTER LOGIN)
# =====================================
@login_required(login_url="/jobcard/login/")
def role_redirect(request):

    profile = UserProfile.objects.filter(user=request.user).first()

    if not profile:
        return redirect("/jobcard/login/")

    # ALWAYS send everyone to the workspace dashboard
    return redirect("jobcard:dashboard_home")


# =====================================
# DEVELOPER USER MANAGEMENT (READ + CREATE)
# =====================================
@role_required("developer")
def user_management(request):

    if request.method == "POST":
        form = UserCreateForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                password=form.cleaned_data["password"]
            )

            UserProfile.objects.create(
                user=user,
                role=form.cleaned_data["role"],
                signature=form.cleaned_data["signature"]
            )

            messages.success(request, "User created successfully!")
            return redirect("jobcard:user_management")

    else:
        form = UserCreateForm()

    users = User.objects.select_related("userprofile").all()

    return render(request, "user_management.html", {
        "form": form,
        "users": users
    })


# =====================================
# EDIT USER
# =====================================
@role_required("developer")
def edit_user(request, user_id):

    user = get_object_or_404(User, id=user_id)
    profile = UserProfile.objects.get(user=user)

    if request.method == "POST":

        new_role = request.POST.get("role")

        # Prevent changing the LAST developer into another role
        if (
            profile.role == "developer"
            and new_role != "developer"
            and UserProfile.objects.filter(role="developer").count() == 1
        ):
            messages.error(
                request,
                "The final Developer account cannot have its role changed."
            )
            return redirect("jobcard:user_management")

        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.save()

        profile.role = new_role
        profile.save()

        messages.success(request, "User updated successfully.")
        return redirect("jobcard:user_management")

    return render(request, "edit_user.html", {
        "user": user,
        "profile": profile,
        "roles": UserProfile.ROLE_CHOICES
    })


# =====================================
# DELETE USER (ENSURE FINAL DEVELOPER PROTECTION)
# =====================================
@role_required("developer")
def delete_user(request, user_id):

    user = get_object_or_404(User, id=user_id)

    profile = UserProfile.objects.filter(user=user).first()

    if profile and profile.role == "developer":

        developers = UserProfile.objects.filter(role="developer").count()

        if developers == 1:
            messages.error(
                request,
                "The final Developer account cannot be deleted."
            )
            return redirect("jobcard:user_management")

    user.delete()

    messages.success(request, "User deleted successfully.")
    return redirect("jobcard:user_management")


# =====================================
# ACTIVATE / DEACTIVATE USER
# =====================================
@role_required("developer")
def toggle_user(request, user_id):

    user = get_object_or_404(User, id=user_id)

    profile = UserProfile.objects.filter(user=user).first()

    # Prevent deactivating the last active developer
    if (
        profile
        and profile.role == "developer"
        and user.is_active
    ):

        active_developers = UserProfile.objects.filter(
            role="developer",
            user__is_active=True
        ).count()

        if active_developers == 1:
            messages.error(
                request,
                "The final active Developer account cannot be deactivated."
            )
            return redirect("jobcard:user_management")

    user.is_active = not user.is_active
    user.save()

    status = "activated" if user.is_active else "deactivated"

    messages.success(request, f"User {status} successfully.")
    return redirect("jobcard:user_management")


# =====================================
# RESET PASSWORD
# =====================================
@role_required("developer")
def reset_password(request, user_id):

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":

        new_password = request.POST.get("password")

        user.password = make_password(new_password)
        user.save()

        messages.success(request, "Password reset successfully")
        return redirect("jobcard:user_management")

    return render(request, "reset_password.html", {
        "user": user
    })





from .models import LineAlert

@login_required(login_url="/jobcard/login/")
def submit_alert(request):

    try:

        if request.method != "POST":
            return JsonResponse(
                {"error": "Invalid request"},
                status=400
            )

        line = request.POST.get("line")
        shift = request.POST.get("shift")
        severity = request.POST.get("severity")
        message = request.POST.get("message")

        if not message:
            return JsonResponse({
                "error": "Please enter a message."
            })

        alert = LineAlert.objects.create(
            operator=request.user,
            line=line,
            shift=shift,
            severity=severity,
            message=message
        )

        return JsonResponse({
            "success": True,
            "alert_id": alert.id
        })

    except Exception as e:

        return JsonResponse({
            "error": str(e)
        }, status=500)



# =====================================
# GET ACTIVE ALERTS
# =====================================
@login_required(login_url="/jobcard/login/")
def get_unresolved_alerts(request):

    alerts = LineAlert.objects.filter(
        status="new"
    ).order_by("-created_at")

    return JsonResponse([
        {
            "id": alert.id,
            "line": alert.line,
            "shift": alert.shift,
            "severity": alert.severity,
            "message": alert.message,
            "operator": alert.operator.username if alert.operator else "Unknown",
            "time": alert.created_at.strftime("%H:%M")
        }
        for alert in alerts
    ], safe=False)


# =====================================
# ACKNOWLEDGE ALERT
# =====================================
@login_required(login_url="/jobcard/login/")
def acknowledge_alert(request, alert_id):

    if request.method != "POST":
        return JsonResponse({
            "success": False,
            "error": "POST request required"
        }, status=400)

    try:
        alert = LineAlert.objects.get(id=alert_id)

        alert.status = "acknowledged"
        alert.acknowledged_by = request.user
        alert.acknowledged_at = timezone.now()

        alert.save()

        return JsonResponse({
            "success": True
        })

    except LineAlert.DoesNotExist:

        return JsonResponse({
            "success": False,
            "error": "Alert not found"
        }, status=404)